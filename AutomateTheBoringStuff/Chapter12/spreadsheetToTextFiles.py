'''
Write a program that performs the tasks of the previous program in reverse order:
The program should open a spreadsheet and write the cells of column A into one text file,
the cells of column B into another text file, and so on.
'''
import openpyxl
from openpyxl.cell import get_column_letter

wb = openpyxl.load_workbook('spreadsheetToTextFile.xlsx')
sheet = wb.active
files = ['file1.txt','file2.txt','file3.txt']

#list to store file contents
lstContent = []
lstValues = []

cols = sheet.max_column

for i in range(cols):
    for v in sheet.columns[i]:
        if v.value != None:
            lstValues.append(v.value)
    lstContent.append(lstValues)
    lstValues = []

#write contents to each separate file
for i in range(len(lstContent)):
    for v in range(len(lstContent[i])):
        f = open(files[i],'a')
        f.write(lstContent[i][v])
        f.write('\n')
        f.close()

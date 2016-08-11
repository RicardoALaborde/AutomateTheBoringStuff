'''
Create a program blankRowInserter.py that takes two integers and a filename string as command line arguments.
Let’s call the first integer N and the second integer M. Starting at row N, the program should insert M blank
rows into the spreadsheet.

For example, when the program is run like this:

python blankRowInserter.py 3 2 myProduce.xlsx

You can write this program by reading in the contents of the spreadsheet.
Then, when writing out the new spreadsheet, use a for loop to copy the first
N lines. For the remaining lines, add M to the row number in the output spreadsheet.
'''

import openpyxl,sys,os
from openpyxl.cell import get_column_letter, column_index_from_string

try:
    #validate we're getting an integer from the given arguments
    int(sys.argv[1])
    int(sys.argv[2])
    #validate we're getting a valid excel workbook
    if os.path.exists(sys.argv[3]):
        wb = openpyxl.load_workbook(sys.argv[3])
    else:
        raise
except ValueError:
    print('Enter an integer for argument 1 and 2.')
    sys.exit()
except:
    print('The given path for the excel sheet is not valid.')
    sys.exit()

wb = openpyxl.load_workbook(sys.argv[3])
sheet = wb.active

rows = sheet.max_row
cols = sheet.max_column

values=[]

for row in range(1,sheet.max_row+1 ):
    for col in range(1,sheet.max_column+1):
            #read from A1 to wherever it ends
            values.append(sheet[get_column_letter(col)+str(row)].value)

wb.remove_sheet(sheet)
wb.create_sheet('Blank Row Inserted')

#separate my list into groups of whatever the biggest col is
lstOfVal = [values[x:x+cols] for x in range(0, len(values), cols)]

sheet = wb.get_sheet_by_name('Blank Row Inserted')

#insert blank rows into list
for ind in range(int(sys.argv[1])):
    lstOfVal.insert(int(sys.argv[2]),[''])

#insert the list into excel
for val in lstOfVal:
    sheet.append(val)

wb.save('blankRowInserter.xlsx')

'''
Write a program to read in the contents of several text files
(you can make the text files yourself) and insert those contents
into a spreadsheet, with one line of text per row. The lines of the
first text file will be in the cells of column A, the lines of the
second text file will be in the cells of column B, and so on.

Use the readlines() File object method to return a list of strings,
one string per line in the file. For the first file, output the first
line to column 1, row 1. The second line should be written to column 1,
row 2, and so on. The next file that is read with readlines() will be written
to column 2, the next file to column 3, and so on.
'''
import openpyxl
from openpyxl.cell import get_column_letter

wb = openpyxl.load_workbook('textFilesToSpreadsheet.xlsx')
sheet = wb.active
files = ['file1.txt','file2.txt','file3.txt']

#list to store file contents
lstContent = []

for i in range(len(files)):
    f = open(files[i],'r')
    #instead of using readlines, im going to use read in order to remove the newlines
    lstContent.append(f.read().splitlines())
    f.close()

for i in range(len(lstContent)):
    for v in range(len(lstContent[i])):
        sheet[str(get_column_letter(i+1))+str(v+1)] = lstContent[i][v]

wb.save('textFilesToSpreadsheet.xlsx')

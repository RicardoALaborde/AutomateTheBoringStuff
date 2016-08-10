'''
Create a program multiplicationTable.py that takes a number N from the command line
and creates an N×N multiplication table in an Excel spreadsheet. For example, when
the program is run like this:

py multiplicationTable.py 6
'''
import openpyxl,sys
from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Font

try:
    #validate we're getting an integer from the given arguments
    int(sys.argv[1])
except:
    print('Enter an integer')
    sys.exit()

workBook=openpyxl.Workbook()
sheet=workBook.active
sheet.title = str(sys.argv[1]) + ' Multiplication Table'
multiply = 1
lstA = []
fontObj = Font(bold=True)
for i in range(int(sys.argv[1])):
    multiply +=1
    #columns
    sheet[get_column_letter(multiply)+'1'].font = fontObj
    sheet[get_column_letter(multiply)+'1']=multiply-1
    lstA.append(multiply-1)
    #rows
    sheet['A'+str(multiply)].font = fontObj
    sheet['A'+str(multiply)]=multiply-1

letter = 2
multiplyIndex=0
matrixLen = int(sys.argv[1])

for i in range(matrixLen):
    row=2
    for m in range(matrixLen):
        cell = get_column_letter(letter)+str(row)
        sheet[cell]=lstA[i]*lstA[multiplyIndex+m]
        row+=1
    letter+=1

workBook.save('MultiplicationTable.xlsx')

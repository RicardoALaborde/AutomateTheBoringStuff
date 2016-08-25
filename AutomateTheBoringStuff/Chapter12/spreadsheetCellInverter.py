'''
Write a program to invert the row and column of the cells in the spreadsheet.
For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa).
This should be done for all cells in the spreadsheet.
'''

import openpyxl
from openpyxl.cell import get_column_letter

wb = openpyxl.load_workbook('spreadsheetCellInverter.xlsx')
sheet = wb.active

rows = sheet.max_row
cols = sheet.max_column

lstCol=[]
lstValues = []

#create the lists that will be inverted
for i in range(cols):
    for cellObj in sheet.columns[i]:
        lstValues.append(cellObj.value)
    lstCol.append(lstValues)
    lstValues = []

#clear the spreadsheet contents
for col in range(cols):
    for cellObj in sheet.columns[col]:
        cellObj.value = None

#write out the lists into the spreadsheet
for i in range(cols):
    for v in range(rows):
        sheet[str(get_column_letter(v+1))+str(i+1)] = lstCol[i][v]


wb.save('spreadsheetCellInverter.xlsx')

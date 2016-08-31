'''
Excel can save a spreadsheet to a CSV file with a few mouse clicks,
but if you had to convert hundreds of Excel files to CSVs, it would
take hours of clicking. Using the openpyxl module from Chapter 12,
write a program that reads all the Excel files in the current working
directory and outputs them as CSV files.

A single Excel file might contain multiple sheets; you’ll have to create
one CSV file per sheet. The filenames of the CSV files should be
<excel filename>_<sheet title>.csv, where <excel filename> is the filename
of the Excel file without the file extension (for example, 'spam_data',
not 'spam_data.xlsx') and <sheet title> is the string from the Worksheet
object’s title variable.
'''

import openpyxl,csv,os

directory = 'excelExample\\'
os.chdir(directory)

lstContent = []
lstValues =[]
lstSpreadSheet = []

for files in os.listdir('.'):
    if files.endswith('.xlsx'):
        wb = openpyxl.load_workbook(files)
        lstSheet = wb.get_sheet_names()
        for sheet in lstSheet:
            sheetName = wb.get_sheet_by_name(sheet)
            cols = sheetName.max_column
            rows = sheetName.max_row
            for i in range(rows):
                for v in sheetName.rows[i]:
                    if v.value != None:
                        lstValues.append(v.value)
                lstContent.append(lstValues)
                lstValues = []
            lstSpreadSheet.append(lstContent)
            lstContent = []
        idx=0
        for a in range(len(lstSpreadSheet)): #spreadsheet
            for b in range(len(lstSpreadSheet[a])): #values
                sheetN = lstSheet[idx]
                csvOpen = open(os.path.splitext(files)[0]+'_'+sheetN+'.csv','a',newline='')
                csvWrite = csv.writer(csvOpen)
                csvWrite.writerow(lstSpreadSheet[a][b])
            csvOpen.close()
            idx+=1
        lstSpreadSheet = []
        idx=0
